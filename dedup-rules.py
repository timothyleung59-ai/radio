from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "选歌约束规则 v2"

header_font = Font(bold=True, color="FFFFFF", size=11)
layer1_fill = PatternFill("solid", fgColor="C0392B")
layer2_fill = PatternFill("solid", fgColor="2980B9")
layer3_fill = PatternFill("solid", fgColor="27AE60")
layer4_fill = PatternFill("solid", fgColor="8E44AD")
header_fill = PatternFill("solid", fgColor="34495E")
wrap = Alignment(wrap_text=True, vertical="top")
border = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)

headers = ["层级", "编号", "规则名称", "检查内容", "数据源 / 触发位置", "违反后果", "代码位置"]
ws.append(headers)
for col, _ in enumerate(headers, 1):
    c = ws.cell(row=1, column=col)
    c.font = header_font
    c.fill = header_fill
    c.alignment = wrap
    c.border = border

rows = [
    # 第 0 层：候选池预过滤（Phase 3，新加）
    ["第0层 候选池预过滤", "0.1", "buildCandidatePool 过滤",
     "构建候选池时已经按 dupId / dupName / 超艺人配额三重过滤；LLM 只能从池子里选 pickIndex",
     "网易云 /simi/song（基于 currentSong）+ DB 高分历史（30天/7天/全量三档）",
     "结构上不可能违反——池里没有 = LLM 不可见",
     "server.js buildCandidatePool 区"],
    ["第0层 候选池预过滤", "0.2", "Pool 模式启用阈值",
     "候选池 ≥ 5 首 → 走 pool 模式（pickIndex 输出）；< 5 首 → 退回 free 模式",
     "candidatePool.length 判断",
     "Pool 模式 attempt=1 命中率应 95%+",
     "server.js usePool 判定"],

    # 第 1 层：服务端硬拒（pool & free 共用）
    ["第1层 服务端硬拒", "1.1", "dupById（网易云 ID 撞车）",
     "AI 选出的歌经 resolveSong 拿到网易云 song_id 后，检查是否在 playedIdSet",
     "DB 最近 80 行 play_history.song_id + 客户端 recent + prefetch queue",
     "Free: continue 重试，最多 3 次；3 次都中 → 走兜底\nPool: 候选池已过滤，理论不会触发",
     "server.js post-resolve dedup"],
    ["第1层 服务端硬拒", "1.2", "dupByName（归一化歌名+艺人撞车）",
     "把歌名+艺人都 normalizeKey 后比对 noRepeatMap",
     "同 1.1，规则应对 Live/Remix/不同专辑 等变体",
     "同 1.1",
     "server.js post-resolve dedup"],
    ["第1层 服务端硬拒", "1.3", "艺人配额（取代旧 5 不同艺人硬封禁）",
     "三档时间窗：最近 10 首 ≤ 1 首 / 最近 30 首 ≤ 2 首 / 最近 30 天 ≤ 5 首；任一超额都拒",
     "recentPlayed (queue + 最近播过) + dbRecent + DB SQL 聚合 30 天",
     "Free: continue 重试\nPool: 池构建时已过滤",
     "server.js isArtistOverQuota"],
    ["第1层 服务端硬拒", "1.4", "网易云搜不到可播 ID",
     "resolveSong 返回 null 或没有 id（搜不到 / VIP / 下架）",
     "网易云 /cloudsearch API",
     "Free: continue 重试\nPool: 池构建时已过滤无 id 候选",
     "server.js resolveSong"],
    ["第1层 服务端硬拒", "1.5", "JSON 解析失败 / pickIndex 越界",
     "Free 模式：extractJsonFromBlocks 拿不到 candidate.song.name\nPool 模式：pickIndex 不是 1-N 整数",
     "Anthropic response.content",
     "continue 重试；3 次都中 → 走兜底",
     "server.js retry loop"],

    # 第 2 层：sysPrompt 硬规则（pool 模式）
    ["第2层 sysPrompt 硬规则 (Pool)", "2P.1", "JSON 严格输出",
     '输出严格 JSON，第一个字符必须是 "{"',
     "pool sysPrompt 规则 1",
     "JSON 解析失败 → 触发 1.5 重试",
     "server.js usePool sysPrompt"],
    ["第2层 sysPrompt 硬规则 (Pool)", "2P.2", "必须从候选池挑 pickIndex",
     "禁止生成池外的歌名；输出对应的 pickIndex（1-N）",
     "pool sysPrompt 规则 2",
     "越界 → 1.5 拒绝重试",
     "server.js usePool sysPrompt"],
    ["第2层 sysPrompt 硬规则 (Pool)", "2P.3", "氛围承接 + 风格匹配",
     "选歌优先承接【上一首（承接基线）】氛围，符合【模式 / 风格 / 语调】",
     "pool sysPrompt 规则 3",
     "AI 自觉遵守，无服务端硬拒",
     "server.js usePool sysPrompt"],
    ["第2层 sysPrompt 硬规则 (Pool)", "2P.4", "探索艺人配额",
     "池子里有 ⭐ 标记的探索艺人，每 3-5 首挑一次他们的",
     "pool sysPrompt 规则 4 + ctx 探索建议",
     "AI 自觉遵守，无服务端硬拒",
     "server.js usePool sysPrompt"],
    ["第2层 sysPrompt 硬规则 (Pool)", "2P.5", "串词指南遵守",
     "串词严格按【DJ 串词指南】里的长度、语调、承接、避免复读各项要求",
     "pool sysPrompt 规则 5（仅 introSpec 非空时）",
     "AI 自觉遵守",
     "server.js usePool sysPrompt"],

    # 第 2 层：sysPrompt 硬规则（free 模式）
    ["第2层 sysPrompt 硬规则 (Free)", "2F.1", "JSON 严格输出 + 单首歌",
     "输出严格 JSON；只挑 1 首歌",
     "free sysPrompt 规则 1, 2",
     "解析失败 → 1.5 重试",
     "server.js free sysPrompt"],
    ["第2层 sysPrompt 硬规则 (Free)", "2F.2", "🚫 歌单禁列",
     "🚫 编号歌单里的歌一首都不能选（含 Live/Remix/翻唱/不同版本）",
     "free sysPrompt 规则 3a + ctx 顶部禁列",
     "AI 偶尔违反 → 1.1/1.2 兜底",
     "server.js free sysPrompt"],
    ["第2层 sysPrompt 硬规则 (Free)", "2F.3", "🚫 艺人禁列",
     "🚫 艺人禁列里的艺人，他们的任何歌都不能选——不要只换歌却保留同艺人",
     "free sysPrompt 规则 3b + ctx 艺人禁列",
     "AI 偶尔违反 → 1.3 兜底",
     "server.js free sysPrompt"],
    ["第2层 sysPrompt 硬规则 (Free)", "2F.4", "风格匹配 + 网易云可达 + 多样性",
     "风格匹配 / 优先网易云能找到 / 每 5 首至少 1 首探索艺人",
     "free sysPrompt 规则 4-6",
     "AI 自觉遵守，无服务端硬拒",
     "server.js free sysPrompt"],
    ["第2层 sysPrompt 硬规则 (Free)", "2F.5", "串词指南遵守",
     "串词严格按【DJ 串词指南】里的长度、语调、承接、避免复读",
     "free sysPrompt 规则 7（仅 introSpec 非空时）",
     "AI 自觉遵守",
     "server.js free sysPrompt"],

    # 第 3 层：retry 累积反馈
    ["第3层 retry 累积反馈", "3.1", "badPicks 累积",
     "每次被拒的尝试都加进 badPicks 数组，下一轮 prompt 看到完整列表",
     "retry 循环内 badPicks.push",
     "AI 一次看到所有失败，更不容易再选同艺人",
     "server.js retry loop"],
    ["第3层 retry 累积反馈", "3.2", "禁艺人列表",
     "把 badPicks 里的所有艺人列出，要求'本轮内绝对不要再选'",
     "retry 循环内 bannedArtists 计算",
     "强压 AI 换全新艺人",
     "server.js retry loop"],

    # 第 4 层：兜底（fallback）
    ["第4层 兜底", "4.1", "DB 高分历史兜底",
     "三档：30 天前 score≥1 / 7 天前 score≥1 / 任意 score≥1，依次找符合 dedup+quota 的歌",
     "play_history GROUP BY song_id ORDER BY RANDOM()",
     "返回兜底歌，不再 404，prefetch 不卡",
     "server.js fallback path"],
    ["第4层 兜底", "4.2", "兜底命中失败",
     "三档都拿不出符合条件的歌（小众用户极少历史）",
     "if !fallback",
     "返 502/404，客户端 prefetch 退避后再试",
     "server.js fallback path"],
]

for r in rows:
    ws.append(r)

last = len(rows) + 1
for row_idx in range(2, last + 1):
    layer = ws.cell(row=row_idx, column=1).value
    if layer.startswith("第0层"):
        fill = PatternFill("solid", fgColor="D35400")
    elif layer.startswith("第1层"):
        fill = layer1_fill
    elif layer.startswith("第2层"):
        fill = layer2_fill
    elif layer.startswith("第3层"):
        fill = layer3_fill
    else:
        fill = layer4_fill
    for col in range(1, 8):
        c = ws.cell(row=row_idx, column=col)
        c.alignment = wrap
        c.border = border
        if col == 1:
            c.fill = fill
            c.font = Font(bold=True, color="FFFFFF")

widths = [22, 8, 30, 50, 38, 32, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

for row_idx in range(2, last + 1):
    ws.row_dimensions[row_idx].height = 70
ws.row_dimensions[1].height = 28

ws.freeze_panes = "A2"

out = r"C:\Users\xiaotim\Documents\Claude\Projects\radio\dedup-rules.xlsx"
wb.save(out)
print(out)
